import tkinter as tk
from tkinter import messagebox as mb, filedialog
from tkinter import ttk
import socket, threading
import json
from pystray import Icon, MenuItem, Menu
from PIL import Image
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os, sys, re
from pathlib import Path
import queue
import time
from datetime import datetime, timedelta



class ITReportForm(tk.Tk):
    def __init__(self, mode, port=None, client_socket=None):
        super().__init__()
        self.mode = mode
        self.set_config_file()
        self.client_socket = client_socket
        self.port = port
        self.data = None
        
        if self.mode == "server":
            self.tray_icon = None
            self.issue_option = "user-specific"
            self.data_ready = threading.Event()
            self.writing_file_path = self.get_writing_file_path()
            self.file_format_config(self.writing_file_path)


        base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
        icon_path = os.path.join(base_path, "sign.ico")
        self.title("IT Report Form")
        self.iconbitmap(icon_path)
        
        if self.mode == "server":
            self.geometry(self._center_window(400, 450))
        else:
            self.geometry(self._center_window(400, 250))


        self.configure_gui(self, 4, 1)

        self.create_widget()
        # if self.mode == "server":
        #     self.start_server_thread()

    def _center_window(self, width, height):
        x = (self.winfo_screenwidth() - width) // 2
        y = (self.winfo_screenheight() - height) // 2
        return f"{width}x{height}+{x}+{y}"
    
    def configure_gui(self, frame, rows, cols):
        # if 
        for row in range(rows):
            frame.rowconfigure(row, weight=1)
        for col in range(cols):
            frame.columnconfigure(col, weight=1)

    def file_format_config(self, writing_file_path):
        current_date = datetime.now().date()
        try:
            if writing_file_path.split(".")[-1] not in ["xlsx", "xlsm"]:
                raise ValueError("Invalid file selected \nMake sure you have selected valid file for saving data")
            workbook = load_workbook(writing_file_path)
            self.remove_default_sheet(workbook)
            sheet = self.creat_new_worksheet(self.writing_file_path, workbook, current_date) or workbook.active
            self.previous_date = self.sheet_date_checker(sheet)
            self.sheet_date_updater(sheet, current_date)
            workbook.save(writing_file_path)

        except Exception as e:
            self.withdraw()
            mb.showerror("File Error", e)
            self.deiconify()
    
    def sheet_date_updater(self, sheet, current_date):
        if self.previous_date:
            try:
                if (current_date - self.previous_date.date()).days != 0:
                    if current_date < self.previous_date.date():
                        raise ValueError("Sorry your current date configuration might be wrong \nPlease set the correct date before using the application")
                    next_row = sheet.max_row + 2
                    sheet[f"A{next_row}"] = current_date
                    sheet[f"A{next_row}"].number_format = "DD-MMM-YYYY"
                    self.previous_date = datetime.combine(current_date, datetime.min.time())
            except Exception as e:
                self.withdraw()
                mb.showerror("Date Error", e)
                self.destroy()
        else:
            next_row = sheet.max_row + 1
            sheet[f"A{next_row}"] = current_date
            sheet[f"A{next_row}"].number_format = "DD-MMM-YYYY"
            self.previous_date = datetime.combine(current_date, datetime.min.time())
            # return next_row

    def create_widget(self):
        self.font = ("helvetica", 10)
        btn_frm = self.widget(parent=self, name="frame", row=3, col=0, sticky="n")
        btn_frm.grid_columnconfigure(2, weight=1)  # Adjust column width
        btn_frm.grid_rowconfigure(1, weight=1) 
        self.configure_gui(btn_frm, 1, 3)
        
        if self.mode == "client":
            self.client_widgets()
            self.user_lbl_frm.grid(row=0, column=0, sticky="s")
            submit_btn = self.widget(parent=btn_frm, name="button", text="Submit", row=0, col=1, command=self.start_client_thread)
            # self.dropdown_frame = ttk.LabelFrame(self, text="General Problem")
            # self.add_dropdown_content()
            self.config(menu=self.create_menu())

        else:
            self.minimize_to_tray()
            self.client_widgets()
            self.server_widgets()
            self.dropdown_frame = ttk.LabelFrame(self, text="General Problem")
            self.add_dropdown_content()
            self.config(menu=self.create_menu())
            self.specific_issue()
        
            submit_btn = ttk.Button(btn_frm, text="Submit", command=self.on_data_submit)
            submit_btn.grid(row=0, column=0, padx=2.5, pady=2.5)
            discard_btn = ttk.Button(btn_frm, text="Discard", command=self.data_discard)
            discard_btn.grid(row=0, column=1, padx=2.5, pady=2.5)
        

    def widget(self, **kwargs):
        widget = None
        parent = kwargs.get("parent")
        font = kwargs.get("font")
        widget_name = kwargs.get("name")
        text = kwargs.get("text")
        btn_command = kwargs.get("command")
        row = kwargs.get("row")
        col = kwargs.get("col")
        padx = kwargs.get("padx")
        pady = kwargs.get("pady")
        sticky = kwargs.get("sticky")
        colspan = kwargs.get("colspan")
        vals = kwargs.get("values")

        if widget_name == "combo":
            widget = ttk.Combobox(parent, font=font, width=30, values=vals)
        if widget_name == "entry":
            widget = ttk.Entry(parent, font=font, width=32)
        if widget_name == "frame":
            widget = ttk.Frame(parent)
        if widget_name == "lframe":
            widget = ttk.LabelFrame(parent, text=text)
        if widget_name == "button":
            widget = ttk.Button(parent, text=text, command=btn_command)
        if widget_name == "label":
            widget = ttk.Label(parent, text=text, font=(*font, "bold"))
        widget.grid(row=row, column=col, padx=5, pady=5, sticky=sticky, columnspan=colspan)
        return widget

    def client_widgets(self):

        self.user_lbl_frm = ttk.LabelFrame(self, text="User Details")
        self.configure_gui(self.user_lbl_frm, 1, 1)

        usernames = self.user_options().get("usernames")
        username_lbl = self.widget(parent=self.user_lbl_frm, name="label", text="Username:", font=self.font, row=0, col=0, sticky="e")
        self.username = self.widget(parent=self.user_lbl_frm, name="combo", font=self.font, row=0, col=1, sticky="w", values=usernames)
        usernames = self.user_options().get("usernames")
        self.username.bind("<KeyRelease>", lambda event: self.combobox_filter(event, self.username, usernames))

        names = self.user_options().get("names")
        name_lbl = self.widget(parent=self.user_lbl_frm, name="label", text="Name:", font=self.font, row=1, col=0, sticky="e")
        self.name = self.widget(parent=self.user_lbl_frm, name="combo", font=self.font, row=1, col=1, sticky="w", values=names)
        names = self.user_options().get("names")
        self.name.bind("<KeyRelease>", lambda event: self.combobox_filter(event, self.name, names))

        # Get current time in AM/PM-hour format
        current_time = time.strftime("%I:%M:%S %p")
        time_lbl = self.widget(parent=self.user_lbl_frm, name="label", text="Time:", font=self.font, row=3, col=0, sticky="e")
        self.time = self.widget(parent=self.user_lbl_frm, name="entry", font=self.font, row=3, col=1, sticky="w")

        
        if self.mode == "client":
            send_to_lbl = self.widget(parent=self.user_lbl_frm, name="label", text="Send to:", font=self.font, row=4, col=0, sticky="e")
            self.send_to = self.widget(parent=self.user_lbl_frm, name="combo", font=self.font, row=4, col=1, sticky="w")
            self.time.delete(0, tk.END)
            self.time.insert(tk.END, current_time)
        

        if self.mode == "server":
            des_lbl = self.widget(parent=self.user_lbl_frm, name="label", text="Description:", font=self.font, row=2, col=0, sticky="e")
            self.des = self.widget(parent=self.user_lbl_frm, name="entry", font=self.font, row=2, col=1, sticky="w")


            self.name["state"] = "disabled"
            self.username["state"] = "disabled"
            # self.des.config(state='readonly')
            # self.time.config(state='readonly')

    def server_widgets(self):
        self.it_lbl_frm = ttk.LabelFrame(self, text="IT Details")

        ip_lbl = self.widget(parent=self.it_lbl_frm, name="label", text="IP address:", font=self.font, row=0, col=0, sticky="e")
        self.ip_address_widget = self.widget(parent=self.it_lbl_frm, name="combo", font=self.font, row=0, col=1, sticky="w")
        self.ip_address_widget["state"] = "disabled"

        it_lbl = self.widget(parent=self.it_lbl_frm, name="label", text="IT Personel:", font=self.font, row=1, col=0, sticky="e")
        it_names = self.user_options().get("it")
        self.it = self.widget(parent=self.it_lbl_frm, name="combo", font=self.font, row=1, col=1, sticky="w", values=it_names)
        self.it.bind("<KeyRelease>", lambda event: self.combobox_filter(event, self.it, it_names))

        issue_lbl = self.widget(parent=self.it_lbl_frm, name="label", text="Issue:", font=self.font, row=2, col=0, sticky="e")
        self.issue = self.widget(parent=self.it_lbl_frm, name="entry", font=self.font, row=2, col=1, sticky="w")

        duration_lbl = self.widget(parent=self.it_lbl_frm, name="label", text="Duration:", font=self.font, row=3, col=0, sticky="e")
        self.duration = self.widget(parent=self.it_lbl_frm, name="entry", font=self.font, row=3, col=1, sticky="w")


    def create_menu(self):

        menu = tk.Menu(self)

        preference_menu = tk.Menu(menu, tearoff=0)
        if self.mode == "server":
            option_menu = tk.Menu(menu, tearoff=0)
            option_menu.add_command(label="Agent Specific", command=self.specific_issue)
            option_menu.add_command(label="Mass Issue", command=self.general_issue)
            preference_menu.add_command(label="Save To", command=self.set_writing_file_path)


            menu.add_cascade(label="Options", menu=option_menu)

        preference_menu.add_command(label="Read from", command=self.set_reading_file_path)

        menu.add_cascade(label="Settings", menu=preference_menu)    
        return menu

    def get_app_folder(self):
        # Cross-platform AppData path resolution
        if os.name == 'nt':  # For Windows
            app_data_path = os.getenv('LOCALAPPDATA')
        else:  # For Linux/macOS
            app_data_path = os.path.expanduser("~/.local/share")

        # Validate if app_data_path is resolved
        if not app_data_path:
            raise EnvironmentError("Unable to locate AppData or equivalent directory.")
        
        # Define the path to your application's folder in AppData
        app_folder = os.path.join(app_data_path, "ReportApp")  # Path to your custom folder
        # Create the folder if it doesn't exist
        os.makedirs(app_folder, exist_ok=True)
        
        return app_folder
    
    def set_writing_file_path(self):
        # Open the file dialog to select a file
        writing_file_path = filedialog.askopenfilename(
            title="Select a file", 
            # initialdir=self.writing_file_path, 
            filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
        )
        self.set_config_file(writing_file_path=writing_file_path)
        self.writing_file_path = writing_file_path
        self.file_format_config(writing_file_path)

    def set_reading_file_path(self):
        # Open the file dialog to select a file
        reading_file_path = filedialog.askopenfilename(
            title="Select a file", 
            # initialdir=self.writing_file_path, 
            filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
        )
        self.set_config_file(reading_file_path=reading_file_path)
        self.user_options()


    def set_config_file(self, writing_file_path=None, reading_file_path=None):

        app_folder = self.get_app_folder()
        config_file = os.path.join(app_folder, "config.json")

        if not os.path.exists(config_file):
            # If the file doesn't exist, create it and write the default data
            # Get the path to the Desktop folder
            desktop_path = Path.home() / 'Desktop'
            config_data = {
                "writing_file_path": os.path.join(desktop_path, "output.xlsx"),  # Default path or data
                "reading_file_path": os.path.join(os.path.dirname(sys.executable), "it_agent.xlsx")
            }
            with open(config_file, "w") as file:
                json.dump(config_data, file, indent=4)
        else:
            if writing_file_path:
                try:
                    with open(config_file, "r") as f:
                        config_data = json.load(f)  # Load existing config data
                        # Update the config with the new file path
                        config_data["writing_file_path"] = writing_file_path

                    with open(config_file, "w") as file:
                        json.dump(config_data, file, indent=4)

                except FileNotFoundError:
                    config_file = {}  # If config.json doesn't exist, create a new dictionary
            
            if reading_file_path:
                try:
                    with open(config_file, "r") as f:
                        config_data = json.load(f)  # Load existing config data
                        # Update the config with the new file path
                        config_data["reading_file_path"] = reading_file_path

                    with open(config_file, "w") as file:
                        json.dump(config_data, file, indent=4)

                except FileNotFoundError:
                    config_file = {}  # If config.json doesn't exist, create a new dictionary
        self.config_file = config_file


    def get_writing_file_path(self):
        try:
            with open(self.config_file, "r") as file:
                config_file = json.load(file)
            return config_file.get("writing_file_path")
        except Exception:
            print(Exception)

    def get_reading_file_path(self):
        try:
            with open(self.config_file, "r") as file:
                config_file = json.load(file)
            return config_file.get("reading_file_path")
        except Exception:
            print({Exception})
        

    def general_issue(self):
        self.dropdown_frame.grid(row=2, column=0, sticky="n", padx=5, pady=5)
        self.user_lbl_frm.grid_remove()
        self.it_lbl_frm.grid_remove()
        self.issue_option = "general"


    def specific_issue(self):
        self.user_lbl_frm.grid(row=0, column=0, sticky="s")
        self.it_lbl_frm.grid(row=1, column=0, sticky="n")
        # self.general_issue_txt.delete("1.0", tk.END)
        self.dropdown_frame.grid_remove()  # Hide the frame
        self.issue_option = "user-specific"



    def add_dropdown_content(self):

        # self.issue_option = "general"
        issue_label = ttk.Label(self.dropdown_frame, text="NOTE: ", font=("helvetica", 10, "bold"))
        issue_label.grid(row=0, column=0, padx=5, pady=5, sticky="ne")
        self.general_issue_txt = tk.Text(self.dropdown_frame, wrap="word", height=5, width=40, font=("helvetica", 10, "bold"))
        self.general_issue_txt.grid(row=0, column=1, padx=5, pady=5)
    # def create_receive_widget(self):

    def start_client_thread(self):
        threading.Thread(target=self.connect_server, args=(self.client_socket,), daemon=True).start()

    def minimize_to_tray(self):
    # Hide the window
        self.withdraw()

        # Create an icon for the tray
        image = Image.open("sign.ico")  # Replace with your icon file
        menu = Menu(MenuItem("Show", self.show_window), MenuItem("Exit", self.exit_app))
        self.tray_icon = Icon("name", image, "Report App", menu)
        self.tray_icon.run()

    # Function to handle showing the Tkinter window
    def show_window(self, icon=None, item=None):
        self.deiconify()
        if self.tray_icon:
            self.tray_icon.stop()
        self.tray_icon = None

    def exit_app(self, icon=None, item=None):
        if self.tray_icon:
            self.tray_icon.stop()
        self.destroy()

    def get_month_range_name(self):
        """
        Generates a string representing the current and next month in the format:
        'CurrentMonth-NextMonth YYYY'
        """
        curr_date = datetime.now()
        # Get the current and next month
        current_month = curr_date.strftime("%B")  # e.g., "October"
        next_month_date = curr_date + timedelta(days=12)  # Jump to the next month from 21th by 21 + 12 = 33
        next_month = next_month_date.strftime("%B")  # e.g., "November"
        
        # Get the year
        year = next_month_date.strftime("%Y")  # Year corresponds to the next month
        
        # Create the sheet name
        sheet_name = f"{current_month}-{next_month} {year}"
        return sheet_name

    def remove_default_sheet(self, workbook):
        # Check if the workbook has only one sheet named "Sheet1"
        if len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet1":
            # Remove "Sheet1"
            sheet_to_remove = workbook["Sheet1"]
            workbook.remove(sheet_to_remove)    


    def creat_new_worksheet(self, writing_file_path, workbook, today):
         # Path to your existing workbook
        if today.day >= 21 or len(workbook.sheetnames) <= 1:
            # Format the sheet name as YYYY-MM-DD
            sheet_name = self.get_month_range_name()
            


            # Check if the sheet already exists to avoid duplication
            if sheet_name not in workbook.sheetnames:
                # Add a new sheet with the specific name
                worksheet = workbook.create_sheet(title=sheet_name)
                workbook.active = workbook.sheetnames.index(worksheet.title)
                # Add the main title
                month = self.get_month_range_name().split("-")
                prev_month = month[0]
                curr_month = month[1].split(" ")[0]
                main_title = f"IT RELATED CASE REPORT AS OF {prev_month.upper()} 21 TO {curr_month.upper()} 20"
                worksheet.merge_cells("A1:I2")  # Merge cells A1 to I1 for the title
                title_cell = worksheet["A1"]
                title_cell.value = main_title
                title_cell.border = self.cell_border()

                # Style the title
                title_cell.font = Font(name="Calibri", size=16, bold=True)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")

                # Add the field headers
                fields = [
                    "Date", "IT", "Username", "Name", "Designation",
                    "Computer IP", "Time", "Resolution Time(min)", "Issue"
                ]
                # Set headers in row 3, starting from column A
                large_cell = [2, 3, 4]
                small_cell = [5, 6, 7]
                medium_cell = [1, 8]
                larger_cell = [9]
                for col, field in enumerate(fields, start=1):
                    cell = worksheet.cell(row=3, column=col, value=field)
                    if col in large_cell:
                        worksheet.column_dimensions[cell.column_letter].width = 28
                    if col in small_cell:
                        worksheet.column_dimensions[cell.column_letter].width = 13
                    if col in medium_cell:
                        worksheet.column_dimensions[cell.column_letter].width = 15
                    if col in larger_cell:
                        worksheet.column_dimensions[cell.column_letter].width = 50

                # Style the headers with background color
                header_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Light blue background
                for col in range(1, len(fields) + 1):  # Loop through all columns in the header row
                    cell = worksheet.cell(row=3, column=col)  # Header is on the second row
                    cell.fill = header_fill
                    cell.font = Font(name="Calibri", size=11, bold=True)  # Make header text bold
                    cell.border = self.cell_border()
                    if col == 8:
                        cell.font = Font(name="calibri", size=9, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align text
                # for col in worksheet.c

            workbook.save(writing_file_path)
            return workbook.active
        return None

    
    def sheet_date_checker(self, sheet):
        for row in range(sheet.max_row, 0, -1):
            cell_value = sheet[f"A{row}"].value
            if cell_value is not None:
                if isinstance(cell_value, datetime):
                    return cell_value
                try:
                    # Attempt to parse string as date
                    return datetime.strptime(cell_value, "%d-%b-%Y").date()
                except (ValueError, TypeError):
                    # Skip invalid entries (like headers or titles)
                    continue
        return None
    

    def data_reset(self):
        if self.issue_option == "user-specific":
            self.username.set("")
            self.name.set("")
            self.des.delete(0, tk.END)
            self.time.delete(0, tk.END)
            self.ip_address_widget.set("")
            self.duration.delete(0, tk.END)
            self.issue.delete(0, tk.END)
            self.it.set("")
        else:
            self.general_issue_txt.delete("1.0", tk.END)

    def cell_border(self):
        thin_border = Side(style='thin', color="000000")
        return Border(left=thin_border, 
                right=thin_border, 
                top=thin_border, 
                bottom=thin_border)

    def data_discard(self):
        if mb.askyesno("Discard Data", "Do you want to proceed?"):
            self.data_reset()
            self.data_ready.set()
    
    def on_data_submit(self):
        current_date = datetime.now()
        workbook = load_workbook(self.writing_file_path)
        sheet = workbook.active
        try:
            if (current_date - self.previous_date).days != 0:
                self.file_format_config(self.writing_file_path)
            else:
                if sheet[f"B{sheet.max_row}"].value is None:
                    next_row = sheet.max_row
                else:
                    next_row = sheet.max_row + 1
        
            value_error = "Seems like you did not fill all the data \nSorry you can not submit with blank field \nAnd also duration should be an integer value"
            if self.issue_option == "user-specific":
                it_name = self.it.get().upper()
                username = self.username.get()
                name = self.name.get().upper()
                des = self.des.get()
                ip_address = self.ip_address_widget.get()
                time = self.time.get()
                duration = int(self.duration.get())
                issue = self.issue.get().capitalize()
                if it_name == "" or issue == "" or duration == "" or username == "":
                    raise ValueError(value_error)
                duration = int(duration)
                data = [it_name, username, name, des, ip_address, time, duration, issue]
                center_align_col = [5, 8]
                
                for col, value in enumerate(data, start=2):
                    date_cell = sheet.cell(row=next_row, column=1)    #selecting date cells
                    date_cell.border = self.cell_border()               #applying borders to date cells               
                    sheet.cell(row=next_row, column=col, value=value)
                    cell = sheet.cell(next_row, col)
                    cell.border = self.cell_border()
                    if col in center_align_col:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                general_issue = self.general_issue_txt.get("1.0", tk.END).upper().strip()
                if general_issue == "":
                    raise ValueError(value_error)
                start_col = 2
                end_col =  9
                data = general_issue
                merge_range = f"{get_column_letter(start_col)}{next_row}:{get_column_letter(end_col)}{next_row}"
                sheet.merge_cells(merge_range)
                sheet.cell(row=next_row, column=start_col, value=data)
            workbook.save(self.writing_file_path)
            next_row = sheet.max_row + 1
            self.data_reset()
            self.specific_issue()
            self.data_ready.set()
            # self.minimize_to_tray()
        except AttributeError:
            mb.showerror("File Error", f"There is issue in writing to the file: {self.writing_file_path} \nClose the file if it is open, then reselect it as saving file from the settings option menu, \nOR simply select another file")
        except ValueError:
            mb.showerror("Value Error", value_error)
        # except TypeError as e:
        #     mb.showerror("Type Error", e)
        except Exception as e:
            mb.showerror("Submission Error", e)
        # finally:

    
    # a mothod for sending data to the server 
    def send_data(self, client_socket):
        # data['ip_address'] = client_socket.laddr[0]
        data = json.dumps(self.update_user_data()).encode("utf-8")
        try:
            client_socket.sendall(data)
            client_socket.close()
            self.destroy()
        except Exception as e:
            print(f"Error: {e}")

    def combobox_filter(self, event, combobox, option_values):
        input_value = combobox.get()
        threading.Thread(target=self.filter_options, args=(input_value, option_values, combobox), daemon=True).start()


    def filter_options(self, typed_text, option_values, combobox):
        # option_values = self.user_options().get("usernames")
        filtered_values = []
        if typed_text == "":
            combobox["values"] = option_values
        else:
            for option in option_values:
                if typed_text.lower() in option.lower():
                    filtered_values.append(option)
            self.after(0, lambda: self.update_combobox(typed_text, filtered_values, combobox))

    def update_combobox(self, typed_text, values, combobox):
        combobox["values"] = values
        combobox.delete(0, tk.END)            # Clear the entry
        combobox.insert(0, typed_text)        # Restore typed text
        combobox.icursor(len(typed_text)) 
        combobox.focus_set()
        # self.after(50, lambda: self.username.event_generate('<Down>'))

    def user_options(self):
        # try:
            workbook = load_workbook(self.get_reading_file_path())
            sheet = workbook["Sheet1"]
            return {
                "usernames": [cell.value for cell in sheet["B"] if cell.value is not None],
                "names": [cell.value for cell in sheet["A"] if cell.value is not None] ,
                "it": [cell.value for cell in sheet["C"] if cell.value is not None] 
            }
        # except:
        #     return {}

    def connect_server(self, client_socket):
      
        addr = self.send_to.get()
        port = self.port
        ipv4_pattern = re.compile(r'^((25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)$')
                                  
        data = self.update_user_data()
        # self.send_data(client_socket, data)
        username = data.get("username")
        name = data.get("name")
        time = data.get("time")
        if username == "" or name == "" or time == "":
            error_des = "Please Make Sure The Required Fields Are Correctly Filled With Correct Information"
            mb.showerror("Error", error_des)
            return 
           
        if bool(ipv4_pattern.match(addr)):
            try:
                if client_socket is None:
                    client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    client_socket.connect((addr, port))
                    self.client_socket = client_socket
                    self.send_data(client_socket)
                else:
                    self.send_data(client_socket)
                    
            except socket.error as e:
                mb.showinfo("Connection", f"{e}")
                return
        else:
            error_des = "Invalid IP address \nIP Range => 0.0.0.0 - 255.255.255.255 \nExample (172.168.1.10)"
            mb.showerror("IP Address Error", error_des)


    def update_user_data(self, data=None):
        self.data = data
        if self.mode == "client":
            return {
                "username": self.username.get(),
                "name": self.name.get(),
                "time": self.time.get()
            }
        else:
            if self.data:


                self.username.configure(values=self.data.get("username"))
                self.username.set(self.data.get("username"))

                self.name.configure(values=self.data.get("name"))
                self.name.set(self.data.get("name"))
                
                #Set description field with default value "CC" ignoring any value of des received
                self.des.delete(0, tk.END)
                self.des.insert(tk.END, "CC")

                #Set time field with current time ignoring any value received
                self.time.delete(0, tk.END)
                self.time.insert(tk.END, self.data.get("time"))

                self.ip_address_widget.delete(0, tk.END)
                self.ip_address_widget.set(self.data.get("client_ip_addr"))
                self.data_ready.clear()
            self.show_window()


def server_start(server_addr, server_port, data_queue):
    try:
        server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        server_socket.bind((server_addr, server_port))
        server_socket.listen()
        # display_gui(data_queue)
        threading.Thread(target=server_listening, args=(server_socket, data_queue,), daemon=True).start()

    except socket.error as e:
        print(f"Connection error: {e}")


def server_listening(server_socket, data_queue):
    while True:
        try:
            client_socket, addr = server_socket.accept()
            threading.Thread(target=handle_clients, args=(client_socket ,data_queue), daemon=True).start()
        except socket.error as e:
            print(f"Error accepting connection: {e}")


def handle_clients(client_socket, data_queue):
    try:
        data = client_socket.recv(1024).decode("utf-8")
        data = json.loads(data)
        data["client_ip_addr"] = client_socket.getpeername()[0]
        data_queue.put(data)
        
        # client_socket.close()
    except socket.error as e:
        print(f"Socket Error: {e}")
    finally:
        client_socket.close()


def display_gui(queue):
    mode = "server"
    try:
        app = ITReportForm(mode=mode)
        threading.Thread(target=process_gui_queue, args=(app, queue), daemon=True).start()
        app.protocol("WM_DELETE_WINDOW", app.minimize_to_tray)
        app.mainloop()
    except Exception as e:
        print(f"Error: {e}")

def process_gui_queue(app, queue):
    while True:
        try:
            data = queue.get()  # Get data sent from the server
            if data == "STOP":
                break
            # Process data and update GUI
            app.update_user_data(data)
            app.data_ready.wait()

        except Exception as e:
            print(f"GUI update error: {e}")


if __name__ == "__main__":
    hosts = socket.gethostbyname_ex(socket.gethostname())[2]
    required_host = ""
    print(hosts)
    try:
        for host in hosts:
            if "10.6" in host:
                required_host = host
                break
        if required_host == "":
            raise ValueError("Can not find correct ip address range on any interface \nConfigure IP address and Possibly add route to routing table")
    except ValueError as e:
        root = tk.Tk()
        root.withdraw()
        result = mb.showerror("IP Address error", e)
        if result:
            sys.exit()
        root.mainloop()
    port = 12345
    data_queue = queue.Queue()

    # Get the installation directory of the executable
    if getattr(sys, 'frozen', False):
        # When running from the bundled executable, get the installation directory
        app_dir = os.path.dirname(sys.executable)
    else:
        # If running as a script (before packaging), use the current directory
        app_dir = os.path.dirname(os.path.abspath(__file__))
        # Define the path to the config.json file in the installation folder
    config_path = os.path.join(app_dir, 'config.json')

    # Open the config file
    try:
        with open(config_path, 'r') as file:
            config_data = json.load(file)
    except FileNotFoundError:
            print(f"Error: {config_path} not found.")


    while True:
        mode = config_data.get("mode")
        if mode == "client" or mode == "server":
            if mode == "client":
                app = ITReportForm(mode, port)
                app.mainloop()
            else:
                threading.Thread(target=server_start, args=(required_host, port, data_queue), daemon=True).start()
                display_gui(data_queue)
            break